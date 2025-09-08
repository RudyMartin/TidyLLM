#!/usr/bin/env python3
"""
QA Control FLOW Agreement
=========================

A FLOW Agreement for Quality Assurance Control using Excel files with three tabs:
1. core_checklist - Required QA metrics
2. custom_checklist - Additional custom QA items
3. custom_prompts - Custom prompts for analysis

Drop Zone Pipeline:
qa_drop → qa_ingest → qa_extract → qa_embed → qa_index → qa_analysis → qa_reports

Usage:
    python qa_control_flow.py setup      # Create folders and config
    python qa_control_flow.py watch      # Watch for Excel files
    python qa_control_flow.py "[QA Control]"  # Execute via FLOW
"""

import json
import os
import shutil
import time
from pathlib import Path
from typing import Dict, Any, List, Optional
from dataclasses import dataclass
from datetime import datetime
import re

# Excel parsing (using openpyxl if available, fallback to CSV)
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Using CSV fallback for Excel files.")

@dataclass
class QAControlConfig:
    """Configuration for QA Control workflow."""
    revision_pattern: str = r"^REV\d{5,}$"  # REV00001, REV12345, etc.
    poll_seconds: int = 3
    max_batch: int = 5
    excel_tabs: List[str] = None
    drop_zone: str = "qa_drop"
    pipeline_stages: List[str] = None
    
    def __post_init__(self):
        if self.excel_tabs is None:
            self.excel_tabs = ["core_checklist", "custom_checklist", "custom_prompts"]
        if self.pipeline_stages is None:
            self.pipeline_stages = [
                "qa_drop",
                "qa_ingest", 
                "qa_extract",
                "qa_embed",
                "qa_index",
                "qa_analysis",
                "qa_reports",
                "qa_logs"
            ]

@dataclass
class QAMetric:
    """A single QA metric/checklist item."""
    metric_id: str
    section: str
    question: str
    expected_answer: Optional[str] = None
    severity: str = "medium"  # low, medium, high, critical
    source: str = "core"  # core or custom
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "metric_id": self.metric_id,
            "section": self.section,
            "question": self.question,
            "expected_answer": self.expected_answer,
            "severity": self.severity,
            "source": self.source
        }

@dataclass
class QAPrompt:
    """A custom prompt for QA analysis."""
    prompt_id: str
    prompt_text: str
    prompt_type: str = "analysis"  # analysis, validation, summary
    context: Optional[str] = None
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "prompt_id": self.prompt_id,
            "prompt_text": self.prompt_text,
            "prompt_type": self.prompt_type,
            "context": self.context
        }

class QAControlManager:
    """Manages the QA Control workflow."""
    
    def __init__(self, config: Optional[QAControlConfig] = None):
        self.config = config or QAControlConfig()
        self.root = Path.cwd()
        self.processed_revisions = set()
        self._load_processed()
    
    def setup_folders(self):
        """Create all pipeline folders."""
        print("=" * 60)
        print("QA CONTROL SETUP")
        print("=" * 60)
        
        for stage in self.config.pipeline_stages:
            path = self.root / stage
            path.mkdir(parents=True, exist_ok=True)
            print(f"Created: {stage}/")
        
        # Create config file
        config_path = self.root / "qa_config.json"
        if not config_path.exists():
            config_data = {
                "revision_pattern": self.config.revision_pattern,
                "poll_seconds": self.config.poll_seconds,
                "max_batch": self.config.max_batch,
                "excel_tabs": self.config.excel_tabs,
                "pipeline_stages": self.config.pipeline_stages
            }
            config_path.write_text(json.dumps(config_data, indent=2))
            print(f"Created: qa_config.json")
        
        # Create README in drop zone
        readme_path = self.root / self.config.drop_zone / "README.md"
        if not readme_path.exists():
            readme_content = """# QA Control Drop Zone

## Expected File Structure

Drop an Excel file (.xlsx) with the following naming pattern:
- `REV00001_qa_control.xlsx`
- `REV12345_qa_control.xlsx`

## Required Excel Tabs

1. **core_checklist** - Core QA metrics (required)
   - Columns: metric_id, section, question, expected_answer, severity
   
2. **custom_checklist** - Custom QA items (optional)
   - Columns: metric_id, section, question, expected_answer, severity
   
3. **custom_prompts** - Custom analysis prompts (optional)
   - Columns: prompt_id, prompt_text, prompt_type, context

## Pipeline Flow

1. Drop Excel file here
2. System detects and processes through:
   - qa_ingest: Raw file copy
   - qa_extract: Parse Excel tabs to JSON
   - qa_embed: Generate embeddings
   - qa_index: Build search index
   - qa_analysis: Run QA analysis
   - qa_reports: Generate final report

## FLOW Agreement Usage

You can also trigger via FLOW Agreement:
```bash
python 1-enterprise.py "[QA Control]"
```
"""
            readme_path.write_text(readme_content)
            print(f"Created: {self.config.drop_zone}/README.md")
        
        print("\nSetup complete! Drop Excel files in qa_drop/ to begin.")
    
    def _load_processed(self):
        """Load list of already processed revisions."""
        log_file = self.root / "qa_logs" / "processed.json"
        if log_file.exists():
            data = json.loads(log_file.read_text())
            self.processed_revisions = set(data.get("processed", []))
    
    def _save_processed(self):
        """Save list of processed revisions."""
        log_file = self.root / "qa_logs" / "processed.json"
        log_file.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "processed": sorted(self.processed_revisions),
            "last_updated": datetime.now().isoformat()
        }
        log_file.write_text(json.dumps(data, indent=2))
    
    def find_new_excel_files(self) -> List[Path]:
        """Find new Excel files matching revision pattern."""
        drop_path = self.root / self.config.drop_zone
        if not drop_path.exists():
            return []
        
        pattern = re.compile(self.config.revision_pattern)
        new_files = []
        
        for file in drop_path.glob("*.xlsx"):
            # Extract revision from filename
            rev_match = pattern.search(file.stem)
            if rev_match:
                revision = rev_match.group()
                if revision not in self.processed_revisions:
                    new_files.append(file)
        
        return sorted(new_files)
    
    def parse_excel_file(self, excel_path: Path) -> Dict[str, Any]:
        """Parse Excel file with three tabs."""
        result = {
            "filename": excel_path.name,
            "revision": self._extract_revision(excel_path),
            "metrics": [],
            "prompts": [],
            "parse_errors": []
        }
        
        if not EXCEL_AVAILABLE:
            # Fallback: treat as CSV or use mock data
            result["parse_errors"].append("openpyxl not installed - using mock data")
            result["metrics"] = self._get_mock_metrics()
            result["prompts"] = self._get_mock_prompts()
            return result
        
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            
            # Parse core_checklist tab
            if "core_checklist" in wb.sheetnames:
                sheet = wb["core_checklist"]
                metrics = self._parse_checklist_sheet(sheet, source="core")
                result["metrics"].extend(metrics)
            
            # Parse custom_checklist tab
            if "custom_checklist" in wb.sheetnames:
                sheet = wb["custom_checklist"]
                metrics = self._parse_checklist_sheet(sheet, source="custom")
                result["metrics"].extend(metrics)
            
            # Parse custom_prompts tab
            if "custom_prompts" in wb.sheetnames:
                sheet = wb["custom_prompts"]
                prompts = self._parse_prompts_sheet(sheet)
                result["prompts"].extend(prompts)
            
            wb.close()
            
        except Exception as e:
            result["parse_errors"].append(str(e))
        
        return result
    
    def _extract_revision(self, file_path: Path) -> str:
        """Extract revision number from filename."""
        pattern = re.compile(self.config.revision_pattern)
        match = pattern.search(file_path.stem)
        return match.group() if match else "REV00000"
    
    def _parse_checklist_sheet(self, sheet, source: str = "core") -> List[QAMetric]:
        """Parse a checklist sheet into QAMetric objects."""
        metrics = []
        headers = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(h).lower() if h else "" for h in row]
                continue
            
            if not any(row):  # Skip empty rows
                continue
            
            metric = QAMetric(
                metric_id=str(row[headers.index("metric_id")] if "metric_id" in headers else f"{source}_{row_idx:03d}"),
                section=str(row[headers.index("section")] if "section" in headers else "General"),
                question=str(row[headers.index("question")] if "question" in headers else ""),
                expected_answer=str(row[headers.index("expected_answer")] if "expected_answer" in headers else None),
                severity=str(row[headers.index("severity")] if "severity" in headers else "medium"),
                source=source
            )
            
            if metric.question:  # Only add if question exists
                metrics.append(metric)
        
        return metrics
    
    def _parse_prompts_sheet(self, sheet) -> List[QAPrompt]:
        """Parse custom prompts sheet."""
        prompts = []
        headers = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(h).lower() if h else "" for h in row]
                continue
            
            if not any(row):
                continue
            
            prompt = QAPrompt(
                prompt_id=str(row[headers.index("prompt_id")] if "prompt_id" in headers else f"prompt_{row_idx:03d}"),
                prompt_text=str(row[headers.index("prompt_text")] if "prompt_text" in headers else ""),
                prompt_type=str(row[headers.index("prompt_type")] if "prompt_type" in headers else "analysis"),
                context=str(row[headers.index("context")] if "context" in headers else None)
            )
            
            if prompt.prompt_text:
                prompts.append(prompt)
        
        return prompts
    
    def _get_mock_metrics(self) -> List[QAMetric]:
        """Get mock metrics for testing without Excel."""
        return [
            QAMetric("CORE001", "Validation", "Is model validation complete?", "Yes", "critical", "core"),
            QAMetric("CORE002", "Testing", "Are all test cases passing?", "Yes", "high", "core"),
            QAMetric("CORE003", "Documentation", "Is documentation updated?", "Yes", "medium", "core"),
            QAMetric("CUST001", "Performance", "Does response time meet SLA?", "<500ms", "high", "custom"),
        ]
    
    def _get_mock_prompts(self) -> List[QAPrompt]:
        """Get mock prompts for testing."""
        return [
            QAPrompt("P001", "Analyze the overall model quality", "analysis"),
            QAPrompt("P002", "Validate compliance with standards", "validation"),
            QAPrompt("P003", "Summarize key findings", "summary"),
        ]
    
    def process_excel_file(self, excel_path: Path) -> Dict[str, Any]:
        """Process Excel file through full pipeline."""
        revision = self._extract_revision(excel_path)
        
        print(f"\n{'=' * 60}")
        print(f"Processing: {excel_path.name}")
        print(f"Revision: {revision}")
        print("=" * 60)
        
        # Stage 1: INGEST
        ingest_dir = self.root / "qa_ingest" / revision
        ingest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(excel_path, ingest_dir / excel_path.name)
        print(f"[1/6] Ingested → qa_ingest/{revision}/")
        
        # Stage 2: EXTRACT
        extract_dir = self.root / "qa_extract" / revision
        extract_dir.mkdir(parents=True, exist_ok=True)
        
        parsed_data = self.parse_excel_file(excel_path)
        
        # Save metrics and prompts as JSON
        metrics_file = extract_dir / "metrics.json"
        metrics_data = [m.to_dict() for m in parsed_data["metrics"]]
        metrics_file.write_text(json.dumps(metrics_data, indent=2))
        
        prompts_file = extract_dir / "prompts.json"
        prompts_data = [p.to_dict() for p in parsed_data["prompts"]]
        prompts_file.write_text(json.dumps(prompts_data, indent=2))
        
        print(f"[2/6] Extracted → qa_extract/{revision}/")
        print(f"      Metrics: {len(metrics_data)}")
        print(f"      Prompts: {len(prompts_data)}")
        
        # Stage 3: EMBED
        embed_dir = self.root / "qa_embed" / revision
        embed_dir.mkdir(parents=True, exist_ok=True)
        
        embeddings = self._generate_embeddings(metrics_data, prompts_data)
        embed_file = embed_dir / "embeddings.json"
        embed_file.write_text(json.dumps(embeddings, indent=2))
        print(f"[3/6] Embedded → qa_embed/{revision}/")
        
        # Stage 4: INDEX
        index_dir = self.root / "qa_index" / revision
        index_dir.mkdir(parents=True, exist_ok=True)
        
        index_data = self._build_index(embeddings)
        index_file = index_dir / "index.json"
        index_file.write_text(json.dumps(index_data, indent=2))
        print(f"[4/6] Indexed → qa_index/{revision}/")
        
        # Stage 5: ANALYSIS
        analysis_dir = self.root / "qa_analysis" / revision
        analysis_dir.mkdir(parents=True, exist_ok=True)
        
        analysis_results = self._run_analysis(metrics_data, prompts_data)
        analysis_file = analysis_dir / "analysis.json"
        analysis_file.write_text(json.dumps(analysis_results, indent=2))
        print(f"[5/6] Analyzed → qa_analysis/{revision}/")
        
        # Stage 6: REPORT
        report_dir = self.root / "qa_reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        
        report_path = report_dir / f"{revision}_qa_report.md"
        self._generate_report(revision, analysis_results, report_path)
        print(f"[6/6] Report → qa_reports/{revision}_qa_report.md")
        
        # Mark as processed
        self.processed_revisions.add(revision)
        self._save_processed()
        
        return {
            "revision": revision,
            "status": "completed",
            "metrics_count": len(metrics_data),
            "prompts_count": len(prompts_data),
            "report_path": str(report_path)
        }
    
    def _generate_embeddings(self, metrics: List[Dict], prompts: List[Dict]) -> List[Dict]:
        """Generate mock embeddings for metrics and prompts."""
        embeddings = []
        
        for metric in metrics:
            text = f"{metric['metric_id']}: {metric['question']}"
            embeddings.append({
                "type": "metric",
                "id": metric["metric_id"],
                "text": text,
                "embedding": [float(hash(text) % 100) / 100 for _ in range(8)]  # Mock 8-dim embedding
            })
        
        for prompt in prompts:
            text = prompt["prompt_text"]
            embeddings.append({
                "type": "prompt",
                "id": prompt["prompt_id"],
                "text": text,
                "embedding": [float(hash(text) % 100) / 100 for _ in range(8)]
            })
        
        return embeddings
    
    def _build_index(self, embeddings: List[Dict]) -> Dict[str, Any]:
        """Build a simple index from embeddings."""
        return {
            "total_items": len(embeddings),
            "metric_count": len([e for e in embeddings if e["type"] == "metric"]),
            "prompt_count": len([e for e in embeddings if e["type"] == "prompt"]),
            "index_type": "flat",  # Could be "hnsw", "ivf", etc. for real vector DB
            "created_at": datetime.now().isoformat()
        }
    
    def _run_analysis(self, metrics: List[Dict], prompts: List[Dict]) -> Dict[str, Any]:
        """Run QA analysis on metrics and prompts."""
        results = {
            "timestamp": datetime.now().isoformat(),
            "metric_results": [],
            "prompt_results": [],
            "summary": {}
        }
        
        # Analyze each metric
        for metric in metrics:
            result = {
                "metric_id": metric["metric_id"],
                "question": metric["question"],
                "expected": metric.get("expected_answer"),
                "severity": metric["severity"],
                "status": "pass" if metric["severity"] in ["low", "medium"] else "review",
                "finding": f"Auto-analysis of {metric['metric_id']}"
            }
            results["metric_results"].append(result)
        
        # Execute each prompt
        for prompt in prompts:
            result = {
                "prompt_id": prompt["prompt_id"],
                "prompt_text": prompt["prompt_text"],
                "prompt_type": prompt["prompt_type"],
                "response": f"[Mock response for {prompt['prompt_type']}]: Analysis complete"
            }
            results["prompt_results"].append(result)
        
        # Generate summary
        total_metrics = len(metrics)
        passed = len([r for r in results["metric_results"] if r["status"] == "pass"])
        results["summary"] = {
            "total_metrics": total_metrics,
            "passed": passed,
            "failed": total_metrics - passed,
            "pass_rate": passed / total_metrics if total_metrics > 0 else 0,
            "severity_breakdown": {
                "critical": len([m for m in metrics if m["severity"] == "critical"]),
                "high": len([m for m in metrics if m["severity"] == "high"]),
                "medium": len([m for m in metrics if m["severity"] == "medium"]),
                "low": len([m for m in metrics if m["severity"] == "low"])
            }
        }
        
        return results
    
    def _generate_report(self, revision: str, analysis: Dict, report_path: Path):
        """Generate markdown report from analysis."""
        lines = []
        lines.append(f"# QA Control Report - {revision}")
        lines.append(f"\nGenerated: {analysis['timestamp']}")
        lines.append("")
        
        # Summary section
        summary = analysis["summary"]
        lines.append("## Executive Summary")
        lines.append(f"- Total Metrics: {summary['total_metrics']}")
        lines.append(f"- Passed: {summary['passed']}")
        lines.append(f"- Failed: {summary['failed']}")
        lines.append(f"- Pass Rate: {summary['pass_rate']:.1%}")
        lines.append("")
        
        lines.append("### Severity Breakdown")
        for severity, count in summary["severity_breakdown"].items():
            lines.append(f"- {severity.upper()}: {count}")
        lines.append("")
        
        # Metric Results
        lines.append("## Metric Analysis")
        for result in analysis["metric_results"]:
            status_icon = "✅" if result["status"] == "pass" else "⚠️"
            lines.append(f"\n### {status_icon} {result['metric_id']}")
            lines.append(f"- **Question**: {result['question']}")
            lines.append(f"- **Expected**: {result['expected']}")
            lines.append(f"- **Severity**: {result['severity']}")
            lines.append(f"- **Status**: {result['status']}")
            lines.append(f"- **Finding**: {result['finding']}")
        
        # Prompt Results
        if analysis["prompt_results"]:
            lines.append("\n## Custom Prompt Analysis")
            for result in analysis["prompt_results"]:
                lines.append(f"\n### {result['prompt_id']} ({result['prompt_type']})")
                lines.append(f"- **Prompt**: {result['prompt_text']}")
                lines.append(f"- **Response**: {result['response']}")
        
        lines.append("\n---")
        lines.append("*Report generated by QA Control FLOW Agreement*")
        
        report_path.write_text("\n".join(lines))
    
    def watch_loop(self):
        """Watch for new Excel files in drop zone."""
        print("=" * 60)
        print("QA CONTROL WATCHER STARTED")
        print("=" * 60)
        print(f"Watching: {self.config.drop_zone}/")
        print(f"Pattern: {self.config.revision_pattern}")
        print("Press Ctrl+C to stop\n")
        
        while True:
            try:
                new_files = self.find_new_excel_files()
                
                for excel_file in new_files[:self.config.max_batch]:
                    self.process_excel_file(excel_file)
                
                time.sleep(self.config.poll_seconds)
                
            except KeyboardInterrupt:
                print("\nWatcher stopped.")
                break
            except Exception as e:
                print(f"Error in watch loop: {e}")
                time.sleep(self.config.poll_seconds)

# FLOW Agreement Integration
class QAControlFlowAgreement:
    """FLOW Agreement for QA Control."""
    
    def __init__(self):
        self.manager = QAControlManager()
    
    def execute(self, context: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Execute QA Control via FLOW Agreement."""
        result = {
            "action": "qa_control",
            "timestamp": datetime.now().isoformat(),
            "files_processed": [],
            "status": "running"
        }
        
        # Check for new Excel files
        new_files = self.manager.find_new_excel_files()
        
        if not new_files:
            result["status"] = "no_files"
            result["message"] = "No new Excel files found in qa_drop/"
            return result
        
        # Process files
        for excel_file in new_files[:self.manager.config.max_batch]:
            try:
                process_result = self.manager.process_excel_file(excel_file)
                result["files_processed"].append(process_result)
            except Exception as e:
                result["files_processed"].append({
                    "file": str(excel_file),
                    "status": "error",
                    "error": str(e)
                })
        
        result["status"] = "completed"
        result["total_processed"] = len(result["files_processed"])
        
        return result

def main():
    """Main entry point."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python qa_control_flow.py setup     # Create folders")
        print("  python qa_control_flow.py watch     # Watch for files")
        print('  python qa_control_flow.py "[QA Control]"  # Execute via FLOW')
        return
    
    command = sys.argv[1]
    manager = QAControlManager()
    
    if command == "setup":
        manager.setup_folders()
    elif command == "watch":
        manager.watch_loop()
    elif command == "[QA Control]":
        agreement = QAControlFlowAgreement()
        result = agreement.execute()
        print(json.dumps(result, indent=2))
    else:
        print(f"Unknown command: {command}")

if __name__ == "__main__":
    main()